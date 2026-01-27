import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# Constants
COLUMNS_TO_EXPORT = [
    "timestamp", "symbol", "close", "volume_ratio", 
    "show_indicator", "rsi_high", "macd_up", "macd_strong", "buy_signal"
]

# Styling
BOLD_FONT = Font(bold=True)
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


def _log_buy_signal(timestamp, symbol, signal_type):
    """Log buy signal to data/log.txt file, sorted by symbol."""
    try:
        log_dir = "data"
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(log_dir, "log.txt")
        
        # Read existing entries
        existing_entries = []
        if os.path.exists(log_file):
            with open(log_file, "r", encoding="utf-8") as f:
                existing_entries = f.readlines()
        
        # Add new entry
        new_entry = f"{timestamp} - {symbol} - {signal_type}\n"
        existing_entries.append(new_entry)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_entries = []
        for entry in existing_entries:
            if entry not in seen:
                seen.add(entry)
                unique_entries.append(entry)
        existing_entries = unique_entries
        
        # Sort by symbol (second part between " - ")
        def get_symbol(line):
            try:
                parts = line.split(" - ")
                return parts[1].strip() if len(parts) > 1 else ""
            except:
                return ""
        
        existing_entries.sort(key=get_symbol)
        
        # Write back sorted entries
        with open(log_file, "w", encoding="utf-8") as f:
            f.writelines(existing_entries)
        
        # Create sorted by date file
        _create_log_sorted_by_date(existing_entries, log_dir)
        
        print(f"📝 Logged: {new_entry.strip()}")
    except Exception as e:
        print(f"❌ Error logging buy signal: {e}")


def _create_log_sorted_by_date(entries, log_dir):
    """Create a new log file sorted by date."""
    try:
        log_by_date_file = os.path.join(log_dir, "log_by_date.txt")
        
        # Sort entries by timestamp (first part before " - ")
        def get_timestamp(line):
            try:
                parts = line.split(" - ")
                return parts[0].strip() if len(parts) > 0 else ""
            except:
                return ""
        
        sorted_entries = sorted(entries, key=get_timestamp, reverse=True)
        
        # Write sorted entries to new file
        with open(log_by_date_file, "w", encoding="utf-8") as f:
            f.writelines(sorted_entries)
        
        print(f"📅 Updated log_by_date.txt with {len(sorted_entries)} entries")
    except Exception as e:
        print(f"❌ Error creating log_by_date: {e}")


def _get_float_value(value, default=0):
    """Safely convert value to float, handling empty strings."""
    return float(value) if value != '' else default


def _calculate_ma_indicator(data_row):
    """Calculate EMA20 and EMA50 trend indicators."""
    close = data_row.get("close")
    ema_20 = data_row.get("ema_20")
    ema_50 = data_row.get("ema_50")
    
    # EMA20 indicator
    if close and ema_20 and float(close) > float(ema_20):
        indicator = "up ema20"
    else:
        indicator = "down ema20"
    
    # EMA50 indicator
    if ema_20 and ema_50 and float(ema_20) > float(ema_50):
        indicator += ", up ema50"
    else:
        indicator += ", down ema50"
    
    return indicator


def _detect_macd_trend(current_row):
    macd = _get_float_value(current_row.get("macd", 0))
    macd_signal = _get_float_value(current_row.get("macd_signal", 0))
    
    if macd > macd_signal:
        return "up" if macd < 0 else "up ++"
    
    return ""


def _detect_macd_strong(current_row):
    """Detect when MACD is negative and absolute value > 3% of close price."""
    try:
        close = _get_float_value(current_row.get("close", 0))
        macd = _get_float_value(current_row.get("macd", 0))
        
        if close > 0 and macd < 0:
            macd_percent = (abs(macd) / close) * 100
            if macd_percent > 3:
                return f"{macd_percent:.2f}%"
        return ""
    except Exception as e:
        print(f"Error in _detect_macd_strong: {e}")
        return ""


def _detect_buy_signal(data_row, previous_5_days=None, previous_data=None, macd_end_signal_sent=False):
    """Detect buy signal based on: up ema20, up ema50, RSI 52-65, vol_ratio > 1, and ALL 5 days ago up ema20.
    Also detect buy signal after macd_strong ends."""
    try:
        # Check EMA conditions
        close = _get_float_value(data_row.get("close", 0))
        open_price = _get_float_value(data_row.get("open", 0))
        ema_20 = _get_float_value(data_row.get("ema_20", 0))
        ema_50 = _get_float_value(data_row.get("ema_50", 0))
        rsi = _get_float_value(data_row.get("rsi14", 0))
        vol_ratio = _get_float_value(data_row.get("volume_ratio", 0))
        
        # NEW: Check if macd_strong ended in last 30 days (1 month) and current conditions met
        # Only check if we haven't sent the signal yet
        if not macd_end_signal_sent and previous_data and len(previous_data) >= 1:
            # Check last 30 days for macd_strong end
            for lookback in range(1, min(31, len(previous_data) + 1)):
                prev_day = previous_data[-lookback] if lookback <= len(previous_data) else None
                if prev_day:
                    prev_macd_strong = prev_day.get("macd_strong", "")
                    # If previous day had macd_strong
                    if prev_macd_strong and str(prev_macd_strong).strip():
                        # Check if current day meets buy conditions: close > ema20 > ema50
                        if close > ema_20 > ema_50:
                            return f"BUY - MACD END"
                        break  # Stop after finding first macd_strong
        
        # # Check if 5 previous days are all up ema20 and up ema50
        # all_5_days_up_ema = False
        # if previous_5_days and len(previous_5_days) >= 5:
        #     all_5_days_up_ema = True
        #     for day_data in previous_5_days:
        #         close_prev = _get_float_value(day_data.get("close", 0))
        #         ema_20_prev = _get_float_value(day_data.get("ema_20", 0))
        #         ema_50_prev = _get_float_value(day_data.get("ema_50", 0))
        #         # Check both up ema20 and up ema50
        #         if not (close_prev > 0 and ema_20_prev > 0 and ema_50_prev > 0 and 
        #                 close_prev > ema_20_prev and ema_20_prev > ema_50_prev):
        #             all_5_days_up_ema = False
        #             break
        
        # Special check: if 5 days all up ema20&50 and current RSI > 80, return CHECK SELL
        if rsi > 80:
            return "CHECK SELL"
        
        # Special check: if price > ema50 and daily gain > 4%, return CHECK
        if close > 0 and ema_50 > 0 and close > ema_50 and open_price > 0:
            daily_gain = ((close - open_price) / open_price) * 100
            if daily_gain > 4:
                if vol_ratio < 1.5:
                    return "CHECK BÁN THÁO"
                return "CHECK GIÁ HỒI MẠNH"
        
        # # Condition 1: up ema20 and up ema50 (must have valid values)
        # if not (close > 0 and ema_20 > 0 and ema_50 > 0 and close > ema_20 and ema_20 > ema_50):
        #     return ""
        
        # # Condition 2: ALL 5 days ago must also be up ema20
        # if previous_5_days and len(previous_5_days) >= 5:
        #     for day_data in previous_5_days:
        #         close_prev = _get_float_value(day_data.get("close", 0))
        #         ema_20_prev = _get_float_value(day_data.get("ema_20", 0))
        #         if not (close_prev > 0 and ema_20_prev > 0 and close_prev > ema_20_prev):
        #             return ""
        
        # # Condition 4: vol_ratio > 1
        # vol_ratio = _get_float_value(data_row.get("volume_ratio", 0))
        # if not (vol_ratio > 1):
        #     return ""
        
        # return "BUY"
    except Exception as e:
        return ""


def _get_column_indices(worksheet):
    """Extract column indices from worksheet headers."""
    headers = [cell.value for cell in worksheet[1]]
    return {
        "show_indicator": headers.index("show_indicator") + 1 if "show_indicator" in headers else None,
        "rsi_high": headers.index("rsi_high") + 1 if "rsi_high" in headers else None,
        "volume_ratio": headers.index("volume_ratio") + 1 if "volume_ratio" in headers else None,
        "macd_up": headers.index("macd_up") + 1 if "macd_up" in headers else None,
        "macd_strong": headers.index("macd_strong") + 1 if "macd_strong" in headers else None,
        "buy_signal": headers.index("buy_signal") + 1 if "buy_signal" in headers else None,
    }


def _format_show_indicator_cell(cell):
    """Format show_indicator cell if contains 'up ema20'."""
    if cell.value and "up ema20" in str(cell.value).lower():
        cell.font = BOLD_FONT


def _format_rsi_cell(cell):
    """Format RSI cell based on value thresholds."""
    try:
        rsi_value = float(cell.value) if cell.value else 0
        if rsi_value > 75:
            cell.font = BOLD_FONT
            cell.fill = RED_FILL
        elif rsi_value > 60:
            cell.font = BOLD_FONT
    except (ValueError, TypeError):
        pass


def _format_volume_ratio_cell(cell):
    """Format volume_ratio cell if >= 1.2."""
    try:
        vol_ratio = float(cell.value) if cell.value else 0
        if vol_ratio >= 1.2:
            cell.font = BOLD_FONT
            cell.fill = RED_FILL
    except (ValueError, TypeError):
        pass


def _format_macd_up_cell(cell):
    """Format macd_up cell: red fill for 'up ++', bold for 'up'."""
    if cell.value:
        if "up ++" in str(cell.value):
            cell.fill = RED_FILL
            cell.font = BOLD_FONT
        elif "up" in str(cell.value):
            cell.font = BOLD_FONT


def _format_buy_signal_cell(cell):
    """Format buy_signal cell: green fill and bold for BUY, orange for CHECK, red for CHECK SELL."""
    if cell.value:
        value_upper = str(cell.value).upper()
        if value_upper == "BUY" or "BUY - MACD END" in value_upper:
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            cell.font = BOLD_FONT
        elif value_upper == "CHECK GIÁ HỒI MẠNH":
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            cell.font = BOLD_FONT
        elif "CHECK BÁN THÁO" in value_upper:
            cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
            cell.font = BOLD_FONT
        elif "CHECK SELL" in value_upper:
            cell.fill = RED_FILL
            cell.font = BOLD_FONT


def _format_macd_strong_cell(cell):
    """Format macd_strong cell: yellow fill and bold when value exists."""
    if cell.value and str(cell.value).strip():
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        cell.font = BOLD_FONT


def _apply_conditional_formatting(worksheet, column_indices):
    """Apply conditional formatting to all data rows in worksheet."""
    for row in range(2, worksheet.max_row + 1):
        if column_indices["show_indicator"]:
            cell = worksheet.cell(row=row, column=column_indices["show_indicator"])
            _format_show_indicator_cell(cell)
        
        if column_indices["rsi_high"]:
            cell = worksheet.cell(row=row, column=column_indices["rsi_high"])
            _format_rsi_cell(cell)
        
        if column_indices["volume_ratio"]:
            cell = worksheet.cell(row=row, column=column_indices["volume_ratio"])
            _format_volume_ratio_cell(cell)
        
        if column_indices["macd_up"]:
            cell = worksheet.cell(row=row, column=column_indices["macd_up"])
            _format_macd_up_cell(cell)
        
        if column_indices["macd_strong"]:
            cell = worksheet.cell(row=row, column=column_indices["macd_strong"])
            _format_macd_strong_cell(cell)
        
        if column_indices["buy_signal"]:
            cell = worksheet.cell(row=row, column=column_indices["buy_signal"])
            _format_buy_signal_cell(cell)


def _prepare_dataframe(data):
    """Prepare and filter DataFrame with only necessary columns."""
    df = pd.DataFrame(data)
    columns_to_keep = [col for col in COLUMNS_TO_EXPORT if col in df.columns]
    return df[columns_to_keep]


def _save_to_excel(data, symbol):
    """Save processed data to Excel file with conditional formatting."""
    try:
        output_dir = "data"
        os.makedirs(output_dir, exist_ok=True)
        
        df = _prepare_dataframe(data)
        
        filename = f"stock_{symbol}.xlsx"
        output_file = os.path.join(output_dir, filename)
        
        # Save to Excel
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Apply formatting
        wb = load_workbook(output_file)
        ws = wb.active
        
        column_indices = _get_column_indices(ws)
        _apply_conditional_formatting(ws, column_indices)
        
        wb.save(output_file)
        print(f"✅ Saved {symbol} data to {output_file}")
        return output_file
    except Exception as e:
        print(f"❌ Error saving Excel for {symbol}: {e}")
        import traceback
        traceback.print_exc()
        return None


def calculate_stock_indicators(data):
    macd_end_signal_sent = False  # Track if we've sent the MACD END signal
    
    for i, row in enumerate(data):
        row["show_indicator"] = _calculate_ma_indicator(row)
        row["rsi_high"] = row.get("rsi14")
        row["macd_up"] = _detect_macd_trend(row)
        row["macd_strong"] = _detect_macd_strong(row)
        
        # Reset signal flag when macd_strong appears again
        if row["macd_strong"] and str(row["macd_strong"]).strip():
            macd_end_signal_sent = False
        
        # Get data from previous 5 days for buy signal check
        previous_5_days = data[i - 5:i] if i >= 5 else None
        # Get all previous data for macd_strong end detection
        previous_data = data[:i] if i > 0 else None
        row["buy_signal"] = _detect_buy_signal(row, previous_5_days, previous_data, macd_end_signal_sent)
        
        # Mark signal as sent if we just sent it
        if row["buy_signal"] == "BUY - MACD END":
            macd_end_signal_sent = True
            # Log the signal to file
            timestamp = row.get("timestamp", "")
            symbol = row.get("symbol", "")
            _log_buy_signal(timestamp, symbol, "BUY - MACD END")
    
    if data and data[0].get("symbol"):
        symbol = data[0]["symbol"]
        _save_to_excel(data, symbol)
    
    return data
